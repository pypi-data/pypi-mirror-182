#   -*- coding: utf-8 -*-
#   @Time      : 2022/12/25 11:18
#   @Author   : Qianlian WANG
#   @File        : _core.py
#   @Software : PyCharm

# RUE，任何几个人共乘都可以，没有公共交通
# 【输入】
# excel 交通网络信息：OD、路段
# KSP 参数 上限 K
# 出行成本参数：VOT、COI、benchmark price、surge price、杂项成本
# 计算精度

# 【输出】
# OD-Route
# Route-Link
# 路径流量分布

__all__ = ["Net"]

import numpy as np
import pandas as pd
from numpy import *
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment  # 先导入分别可指定单元格字体相关，颜色，和对齐方式的类
import copy     # copy，time都是标准库【诡异，标准库之前写在上面会导致识别不出来这个 import】
import time


def Dijkstra(network, s, d):    # 迪杰斯特拉算法算s-d的最短路径，并返回该路径和代价
    # print("Start Dijstra Path……")
    path = []   # s-d的最短路径
    n = len(network)    # 邻接矩阵维度，即节点个数
    fmax = 9999999
    w = [[0 for i in range(n)]for j in range(n)]    # 邻接矩阵转化成维度矩阵，即0→max
    book = [0 for i in range(n)]    # 是否已经是最小的标记列表
    dis = [fmax for i in range(n)]  # s到其他节点的最小距离
    book[s-1] = 1   # 节点编号从1开始，列表序号从0开始
    midpath = [-1 for i in range(n)]    # 上一跳列表
    u = s-1
    for i in range(n):
        for j in range(n):
            if network[i][j] != 0:
                w[i][j] = network[i][j]     # 0→max
            else:
                w[i][j] = fmax
            if i == s-1 and network[i][j] != 0:     # 直连的节点最小距离就是network[i][j]
                dis[j] = network[i][j]
    for i in range(n-1):    # n-1次遍历，除了s节点
        min = fmax
        for j in range(n):
            if book[j] == 0 and dis[j] < min:   # 如果未遍历且距离最小
                min = dis[j]
                u = j
        book[u] = 1
        for v in range(n):  # u直连的节点遍历一遍
            if dis[v] > dis[u] + w[u][v]:
                dis[v] = dis[u] + w[u][v]
                midpath[v] = u+1    # 上一跳更新
    j = d-1     # j是序号
    path.append(d)  # 因为存储的是上一跳，所以先加入目的节点d，最后倒置
    while midpath[j] != -1:
        path.append(midpath[j])
        j = midpath[j]-1
    path.append(s)
    path.reverse()      # 倒置列表
    # print(path)
    # print(midpath)
    # print(dis)
    return path

def return_path_sum(network, path):
    result = 0
    for i in range(len(path)-1):
        result += network[path[i]-1][path[i+1]-1]
    return result

def add_limit(path, s):     # path=[[[1,3,4,6],5],[[1,3,5,6],7],[[1,2,4,6],8]
    result = []
    for item in path:
        if s in item[0]:
            result.append([s, item[0][item[0].index(s)+1]])
    result = [list(r) for r in list(set([tuple(t) for t in result]))]   # 去重
    return result

def return_shortest_path_with_limit(network, s, d, limit_segment, choice):      # limit_segment=[[3,5],[3,4]]
    mid_net = copy.deepcopy(network)
    for item in limit_segment:
        mid_net[item[0]-1][item[1]-1] = mid_net[item[1]-1][item[0]-1] = 0
    s_index = choice.index(s)
    for point in choice[:s_index]:      # s前面的点是禁用点
        for i in range(len(mid_net)):
            mid_net[point-1][i] = mid_net[i][point-1] = 0
    mid_path = Dijkstra(mid_net, s, d)
    return mid_path

def judge_path_legal(network, path):
    for i in range(len(path)-1):
        if network[path[i]-1][path[i+1]-1] == 0:
            return False
    return True

def k_shortest_path(network, s, d, k):
    k_path = []     # 结果列表
    alter_path = []     # 备选列表
    kk = Dijkstra(network, s, d)
    k_path.append([kk, return_path_sum(network, kk)])
    while True:
        if len(k_path) == k:break
        choice = k_path[-1][0]
        for i in range(len(choice)-1):
            limit_path = [[choice[i], choice[i+1]]]     # 限制选择的路径
            if len(k_path) != 1:
                limit_path.extend(add_limit(k_path[:-1], choice[i]))
            mid_path = choice[:i]
            mid_res = return_shortest_path_with_limit(network, choice[i], d, limit_path, choice)
            if judge_path_legal(network, mid_res):
                mid_path.extend(mid_res)
            else:
                continue
            mid_item = [mid_path, return_path_sum(network, mid_path)]
            if mid_item not in k_path and mid_item not in alter_path:
                alter_path.append(mid_item)
        if len(alter_path) == 0:
            print("总共只有{}条最短路径！".format(len(k_path)))
            return k_path
        alter_path.sort(key = lambda x:x[-1])
        x = alter_path[0][-1]
        y = len(alter_path[0][0])
        u = 0
        for i in range(len(alter_path)):
            if alter_path[i][-1] != x:
                break
            if len(alter_path[i][0]) < y:
                y = len(alter_path[i][0])
                u = i
        k_path.append(alter_path[u])
        alter_path.pop(u)
    # for item in k_path:
    #     print(item)
    return k_path


theta = 1  # Logit公式的参数

class TransNet:
    # 交通网络的类，直接以向量存储 link 和 path
    # route 不加入角色前的 路径
    # path 加入角色前的 路径

    ######定义交通网络类中的基本属性######
    Link_Num = 0  # 路段数量，76条路段
    Capacity = 0  # 路段容量
    FFT = 0  # 路段自由流时间
    OD_Num = 0  # OD对数量，528
    Route_Num = []  # 每个OD对的路径数量
    Route_Start = []    # 每个OD对的路径开始
    Route_all = 0   # 网络中 Route 的总数量
    OD_Route = 0    # OD-Route 转换矩阵
    Route_Link = 0      # Route-Link 转换矩阵
    Role_Num = 0  # 交通网络中的角色数量
    Path_Link = 0  # 路径和路段的转换矩阵
    Link_Path = 0  # 路段和路径的转换矩阵
    Car_Path_Link = 0  # 小汽车的 路径-路段转换矩阵
    OD_demand = []  # 获取各个OD对的出行需求


    ######定义交通网络类中的类函数######
    def InitialNet(self, filename, rider_Ca, K):  # 类的初始化操作，类似于构造函数，构造出一个空网络
        # rider_Ca 车内允许搭载的乘客数量上限
        if rider_Ca >= 1:
            self.Role_Num = 1 + 2*rider_Ca  # 获取角色的数量
        else:
            print("Do not fulfill RUE principle because of no rider role.")

        # 在 ”Link“ 中获取路段的自由流时间、容量、长度，形成列向量
        FFT = pd.read_excel(filename, sheet_name="Link", usecols=[3], skiprows=0)  # 路段的自由流时间
        self.FFT = FFT.values
        Cap = pd.read_excel(filename, sheet_name="Link", usecols=[4], skiprows=0)  # 路段的容量
        self.Capacity = Cap.values
        self.Link_Num = self.FFT.shape[0]    # 路段数量

        # 在 "OD" 中获取网络的OD对信息
        OD_demand = pd.read_excel(filename, sheet_name="OD", usecols=[3], skiprows=0)  # 用户出行需求
        self.OD_Demand = OD_demand.values
        self.OD_Num = self.OD_Demand.shape[0]  # OD对数量 = 528

        link = pd.read_excel(filename, sheet_name="Link", usecols=range(1, 4), skiprows=0)  # 路段端点信息
        link = link.values
        temp_nodelist = []
        for i in range(self.Link_Num):
            for j in range(2):
                temp_nodelist.append(link[i][j])
        nodelist = list(set(temp_nodelist))  # 返回数组中不重复的数
        node_num = len(nodelist)   # 节点数量

        network = np.zeros([node_num, node_num])  # 建立空矩阵
        for i in range(self.Link_Num):
            start_point = link[i, 0] - 1  # 起点的索引
            end_point = link[i, 1] - 1  # 终点的索引
            network[start_point, end_point] = link[i, 2]  # 将连通处置为 自由流阻抗

        OD = pd.read_excel(filename, sheet_name="OD", usecols=range(1, 3), skiprows=0)  # OD对的起终点
        OD = OD.values
        OD_route_list = []  # 每个OD对的前K条最短路
        # 为每个OD对生成最多 K 条路径
        for od_i in range(self.OD_Num):
            s = OD[od_i, 0]  # 起点
            d = OD[od_i, 1]  # 终点
            for k_j in range(K-1):  # k_j = 0, 1, ..., K-2
                k_path = k_shortest_path(network, s, d, K-k_j)    # K-k_j = K, K-1, ..., 2
                if k_path[-1] != k_path[-2]:    # 如果最后两条路径不相同，则该OD对筛选 K-i 条路径
                    break
            if k_j == K-2 and k_path[-1] == k_path[-2]:     # 如果 K=2，且两条路径相同，则设置 K=1
                k_path = k_shortest_path(network, s, d, 1)
            OD_route_list.append(k_path)

        for od_i in range(self.OD_Num):
            k_path = OD_route_list[od_i]
            self.Route_Num.append(len(k_path))   # 记录每个OD对的路径数量 len(k_path)

        self.Route_all = int(sum(self.Route_Num))     # 路径总数量
        self.OD_Route = np.zeros([self.OD_Num, self.Route_all])  # 生成 OD_Route 与 Route_Link 转换矩阵
        self.Route_Link = np.zeros([self.Route_all, self.Link_Num])
        od_route_start = 0
        for i in range(self.OD_Num):
            if i > 0:
                od_route_start += self.Route_Num[i - 1]
            self.Route_Start.append(od_route_start)     # 各OD对路径的开端
            self.OD_Route[i, od_route_start: od_route_start+self.Route_Num[i]] = 1  # OD_Route 填充
            s = OD[i, 0]  # 起点
            d = OD[i, 1]  # 终点
            for j in range(self.Route_Num[i]):
                temp_path = k_path[j][0]
                for m in range(len(temp_path) - 1):
                    link_s = temp_path[m]  # 路段起点
                    link_d = temp_path[m + 1]  # 路段终点
                    for n in range(self.Link_Num):
                        if link_s == link[n, 0] and link_d == link[n, 1]:
                            self.Route_Link[od_route_start + j, n] = 1  # Route_Link 填充

        self.Path_Link = np.zeros([self.Route_all * self.Role_Num, self.Link_Num], dtype=int)  # （包含角色）路径-路段转换矩阵
        for i in range(self.Route_all):
            for k in range(self.Role_Num):
                self.Path_Link[self.Role_Num * i + k, ] = self.Route_Link[i, ]
        self.Path_Link = mat(self.Path_Link)  # （包含角色）路径-路段转换矩阵，（5280×7）×76
        self.Link_Path = self.Path_Link.T  # （包含角色）路径-路段转换矩阵，76×（5280×7）

        # 计算 司机模式特定的 路径-路段转换矩阵
        temp = mat(np.zeros([1, self.Link_Num]))
        self.Car_Path_Link = copy.deepcopy(self.Path_Link)  # Car：SD，RD
        for od_i in range(self.OD_Num):
            for i in range(self.Route_Num[od_i]):
                for j in range(int((self.Role_Num+1)/2), self.Role_Num):   # SD, RD1, RD2, R1, R2  3: 5
                    self.Car_Path_Link[(self.Route_Start[od_i] + i) * self.Role_Num + j, ] = temp


    def Initial_PathFlow(self):     # 设置初始流量值
        pathflow_0 = mat(np.zeros([self.Route_all * self.Role_Num, 1]))
        for od_i in range(self.OD_Num):
            od_pathflow_0 = mat(np.zeros([self.Route_Num[od_i] * self.Role_Num, 1]))
            for i in range(self.Route_Num[od_i]):   # route
                for j in range(self.Role_Num):      # role
                    od_pathflow_0[self.Role_Num * i + j, 0] = self.OD_Demand[od_i, 0] / (self.Route_Num[od_i] * self.Role_Num)
            #  pathflow_0 将对应OD对的 od_pathflow_0
            pathflow_0[self.Route_Start[od_i]*self.Role_Num: (self.Route_Start[od_i]+self.Route_Num[od_i])*self.Role_Num, ] = od_pathflow_0
        return pathflow_0


    ## 路径流量是 （528×10×7）×1 的列向量，角色排列分别是 SD，RD，R，总共有 self.Role_Num = 1 + 2*rider_Ca
    def LinkTime(self, pathflow):  # 根据传入的路径流量，计算路段流量，并用BPR函数得到路段出行时间
        Car_Link_Flow = self.Car_Path_Link.T * pathflow     # 路段车辆流量
        LinkTime = mat(np.zeros([self.Link_Num, 1]))   # 路段出行时间
        for i in range(self.Link_Num):
            LinkTime[i, 0] = self.FFT[i, 0] * (1 + 0.15 * (Car_Link_Flow[i, 0] / self.Capacity[i, 0])**4)   # BPR函数
        return LinkTime


    def ODFlow(self, pathflow):  # 根据传入的流量，统计OD对间不同角色的流量
        od_flow = mat(np.zeros([self.OD_Num * self.Role_Num, 1]))
        # pathflow，（528×10×7）×1
        for od_i in range(self.OD_Num):  # OD对
            for i in range(self.Role_Num):  # 角色
                for j in range(self.Route_Num[od_i]):   # 路径
                    od_flow[self.Role_Num * od_i + i, 0] += pathflow[(self.Route_Start[od_i] + j) * self.Role_Num + i, 0]
        return od_flow


    def PathTime(self, LinkTime):   # 计算OD对间路径的 出行时间
        path_time = self.Car_Path_Link * LinkTime  # 路径出行时间
        return path_time


    def TimeCost(self, path_time, VOT):  # 根据不同角色的VOT，计算某一OD对路径的 TimeCost
        time_cost = mat(np.zeros([self.Route_all * self.Role_Num, 1]))
        for od_i in range(self.OD_Num):     # OD对
            for i in range(self.Role_Num):      # 角色
                for j in range(self.Route_Num[od_i]):       # 路线 route
                    time_cost[(self.Route_Start[od_i] + j) * self.Role_Num + i, 0] \
                        = VOT[i, 0] * path_time[(self.Route_Start[od_i] + j) * self.Role_Num + i, 0]  # 时间成本 = 时间 × VOT
        return time_cost


    def IncvCost(self, path_time, COI):  # 计算某一OD对路径的 Inconvenience Cost
        incv_cost = mat(np.zeros([self.Route_all * self.Role_Num, 1]))
        for od_i in range(self.OD_Num):     # OD对
            for i in range(self.Role_Num):      # 角色
                for j in range(self.Route_Num[od_i]):       # 路线 route
                    incv_cost[(self.Route_Start[od_i] + j) * self.Role_Num + i, 0] \
                        = COI[i, 0] * path_time[(self.Route_Start[od_i] + j) * self.Role_Num + i, 0]  # 不便成本 = 时间 × COI
        return incv_cost


    # 时长费0.4，里程费0.2，浮动定价系数设为6和2
    def Price(self, path_time, od_flow, benchmark, surge_cof):  # 计算共乘收费中的收费
        # benchmark 根据角色有所变化
        benchmark_price = mat(np.zeros([self.Route_all * self.Role_Num, 1]))   # 基础价格
        for i in range(self.Route_all):
            for j in range(self.Role_Num):
                benchmark_price[self.Role_Num * i + j, 0] = benchmark[j, 0]
        surge_price = mat(np.zeros([self.Route_all * self.Role_Num, 1]))   # 浮动价格
        for od_i in range(self.OD_Num):     # OD对
            for i in range(self.Role_Num):  # 角色
                for j in range(self.Route_Num[od_i]):  # 路线 route
                    surge_price[(self.Route_Start[od_i] + j) * self.Role_Num + i, 0] = surge_cof[i, 0] * od_flow[self.Role_Num * od_i + i, 0]
        Price = benchmark_price + surge_price     # 总的费用 = 基础价格 + 浮动定价
        return Price


    def TotalCost(self, path_time, od_flow, VOT, COI, benchmark, surge_cof, sundry):  # 计算总的出行费用
        other_cost = mat(np.zeros([self.Route_all * self.Role_Num, 1]))
        for od_i in range(self.OD_Num):    # OD对
            for i in range(self.Role_Num):  # 角色
                for j in range(self.Route_Num[od_i]):    # 路线 route
                    other_cost[(self.Route_Start[od_i] + j) * self.Role_Num + i, 0] = sundry[i, 0]
        # 将各项相加费用相加起来
        total_cost = self.TimeCost(path_time, VOT) + self.IncvCost(path_time, COI) + self.Price(path_time, od_flow, benchmark, surge_cof) + other_cost
        return total_cost


    def Multiplier(self, total_cost):   # 计算网络中各个OD对、各个路径的乘子 λ2，λ3
        rider_Ca = int((self.Role_Num - 1) / 2)     # 车内最多能坐的人数
        Lambda = mat(np.zeros([self.Route_all, rider_Ca]))  # 乘子 λ
        for od_i in range(self.OD_Num):   # OD对
            for i in range(rider_Ca):  # 角色 如果 rider_Ca=2，则 N 取值应该是 1, 2，i = 0, 1；分母是 N+1 = i+2
                for j in range(self.Route_Num[od_i]):  # 路线 route
                    # 角色分配 0=SD, 1=RD1, 2=RD2, ..., N=RDN, N+1=R1, N+2=R2, ..., N+N=RN
                    Lambda[self.Route_Start[od_i] + j, i] = (1/(i+2)) * \
                                                            (total_cost[(self.Route_Start[od_i] + j) * self.Role_Num + (rider_Ca + i + 1), 0]
                                                             - total_cost[(self.Route_Start[od_i] + j) * self.Role_Num + (i + 1), 0])
        return Lambda


    def GeneralCost(self, total_cost, Lambda):   # 计算 OD 对之间各个路径的广义出行费用
        general_cost = mat(np.zeros([self.Route_all * self.Role_Num, 1]))
        rider_Ca = int((self.Role_Num - 1) / 2)  # 车内最多能坐的人数
        for od_i in range(self.OD_Num):     # OD对
            for j in range(self.Route_Num[od_i]):   # route 路线
                general_cost[(self.Route_Start[od_i] + j) * self.Role_Num + 0, 0] = total_cost[(self.Route_Start[od_i] + j) * self.Role_Num + 0, 0]
                for i in range(rider_Ca):   # i = 0, 1, 2, 3   N = 1, 2, 3, 4
                    # 角色分配 0=SD, 1=RD1, 2=RD2, ..., N=RDN, N+1=R1, N+2=R2, ..., N+N=RN
                    # ridersharing driver
                    general_cost[(self.Route_Start[od_i] + j) * self.Role_Num + (i + 1), 0] \
                        = total_cost[(self.Route_Start[od_i] + j) * self.Role_Num + (i + 1), 0] + (i + 1) * Lambda[self.Route_Start[od_i] + j, i]
                    # rider
                    general_cost[(self.Route_Start[od_i] + j) * self.Role_Num + (rider_Ca + i + 1), 0] \
                        = total_cost[(self.Route_Start[od_i] + j) * self.Role_Num + (rider_Ca + i + 1), 0] - Lambda[self.Route_Start[od_i] + j, i]
        return general_cost


    def AssignFlow(self, general_cost):   #  根据广义出行成本，分配各个 OD 对之间的路径流量
        pathflow = mat(np.zeros([self.Route_all * self.Role_Num, 1]))
        rider_Ca = int((self.Role_Num - 1) / 2)     # 乘客人数
        for od_i in range(self.OD_Num):
            od_demand = self.OD_Demand[od_i, 0]   # 交通网络中的OD需求量
            # 取出矩阵关于该OD对的所有行
            od_gen_cost = general_cost[self.Route_Start[od_i]*self.Role_Num: (self.Route_Start[od_i]+self.Route_Num[od_i])*self.Role_Num]
            min_gen_cost = min(od_gen_cost)  # [x]
            min_value = min_gen_cost[0]  # x
            # 最小值的位置，可以获得重复元素的多个位置，结果以索引形式展示出来
            min_pos = [i for i, x in enumerate(od_gen_cost) if x == min_value]
            # OD 对全有全无分配
            # RUE 不能直接进行全有全无，还要考虑到 ride-matching 带来的角色关系
            for j in min_pos:
                r_m = j % self.Role_Num   # 取余 0=SD, 1=RD1, 2=RD2, 3=RD3, 4=R1, 5=R2, 6=R3
                if r_m == 0:    # SD
                    pathflow[self.Route_Start[od_i] * self.Role_Num + j, 0] = od_demand
                    break
                if r_m > 0 and r_m <= rider_Ca :  # RD
                    pathflow[self.Route_Start[od_i] * self.Role_Num + j, 0] = od_demand / (r_m + 1)     # RD
                    pathflow[self.Route_Start[od_i] * self.Role_Num + j + rider_Ca, 0] = r_m * od_demand / (r_m + 1)    # R
                    break
        return pathflow


    def CA(self, pathflow_0, VOT, COI, benchmark, surge_cof, sundry, accuracy_0):   # CA算法，迭代不动点问题
        time_1 = time.time()
        LinkTime_0 = self.LinkTime(pathflow_0)
        path_time_0 = self.PathTime(LinkTime_0)
        od_flow_0 = self.ODFlow(pathflow_0)
        total_cost_0 = self.TotalCost(path_time_0, od_flow_0, VOT, COI, benchmark, surge_cof, sundry)
        lambda_0 = self.Multiplier(total_cost_0)
        general_cost_0 = self.GeneralCost(total_cost_0, lambda_0)

        general_cost_old = copy.deepcopy(general_cost_0)
        iter = 0
        while iter < 10**6:
            # 迭代过程
            iter += 1
            stepsize = 1 / iter  # 步长表达式
            pathflow_mid = self.AssignFlow(general_cost_old)  # 路径流量
            LinkTime_mid = self.LinkTime(pathflow_mid)  # 路段时间
            path_time_mid = self.PathTime(LinkTime_mid)  # 路径时间
            od_flow_mid = self.ODFlow(pathflow_mid)   # OD的角色流量
            total_cost_mid = self.TotalCost(path_time_mid, od_flow_mid, VOT, COI, benchmark, surge_cof, sundry)  # 总成本
            lambda_mid = self.Multiplier(total_cost_mid)  # 乘子
            general_cost_mid = self.GeneralCost(total_cost_mid, lambda_mid)   # 广义出行成本
            general_cost_new = general_cost_old + stepsize * (general_cost_mid - general_cost_old)  # 不动点问题计算公式

            # 检查收敛条件
            temp_old = 0
            temp_dif = 0
            temp_new = 0
            for i in range(self.Route_all * self.Role_Num):
                temp_old += general_cost_old[i, 0] ** 2
                temp_new += general_cost_new[i, 0] ** 2
                temp_dif += (general_cost_new[i, 0] - general_cost_old[i, 0]) ** 2
            # print("old = ", temp_old, "  new = ", temp_new, "  mid = ", temp_mid, "  dif = ", temp_dif)
            # accuracy = temp_dif ** 0.5  # 绝对的收敛条件
            accuracy = (temp_dif / temp_old) ** 0.5  # 相对的收敛条件
            if accuracy <= accuracy_0:  # 满足收敛条件，停止迭代 10 * (-4)
                print(accuracy)
                print("old = ", temp_old, "  new = ", temp_new, "  dif = ", temp_dif)
                break
            else:
                general_cost_old = copy.deepcopy(general_cost_new)  # 更新广义出行成本，继续迭代

        pathflow_new = self.AssignFlow(general_cost_new)  # 由结果广义出行成本，生成结果路径流量
        print("CA迭代次数 = ", iter)
        time_2 = time.time()
        print("CA总共用时：", time_2 - time_1, "s = ", (time_2 - time_1) / 60, "min")
        return [pathflow_new, general_cost_new]


    def MSA(self, pathflow_0, VOT, COI, benchmark, surge_cof, sundry, accuracy_0):   # CA算法，迭代不动点问题
        time_1 = time.time()
        #  由初始流量，计算初始的广义出行成本
        LinkTime_0 = self.LinkTime(pathflow_0)
        path_time_0 = self.PathTime(LinkTime_0)
        od_flow_0 = self.ODFlow(pathflow_0)
        total_cost_0 = self.TotalCost(path_time_0, od_flow_0, VOT, COI, benchmark, surge_cof, sundry)
        lambda_0 = self.Multiplier(total_cost_0)
        general_cost_0 = self.GeneralCost(total_cost_0, lambda_0)

        general_cost_old = copy.deepcopy(general_cost_0)
        pathflow_old = copy.deepcopy(pathflow_0)
        iter = 0
        while iter < 10**6:
            # 迭代过程
            iter += 1
            stepsize = 1 / iter  # 步长表达式
            pathflow_mid = self.AssignFlow(general_cost_old)  # 路径流量（这个应该是符合流量关系的？）
            pathflow_new = pathflow_old + stepsize * (pathflow_mid - pathflow_old)      # MSA 不动点问题计算公式

            # 检查收敛条件
            temp_old = 0
            temp_dif = 0
            temp_new = 0
            for i in range(self.Route_all * self.Role_Num):
                temp_old += pathflow_old[i, 0] ** 2
                temp_new += pathflow_new[i, 0] ** 2
                temp_dif += (pathflow_new[i, 0] - pathflow_old[i, 0]) ** 2
            # print("old = ", temp_old, "  new = ", temp_new, "  mid = ", temp_mid, "  dif = ", temp_dif)
            accuracy = (temp_dif / temp_old) ** 0.5     # 相对的收敛条件
            # accuracy = temp_dif ** 0.5  # 绝对的收敛条件
            # print("accuracy = ", accuracy)
            if accuracy <= accuracy_0:  # 满足收敛条件，停止迭代 10 * (-4)
                print(accuracy)
                print("old = ", temp_old, "  new = ", temp_new,  "  dif = ", temp_dif)
                # 基于结果流量，再计算一次 广义出行成本
                LinkTime_new = self.LinkTime(pathflow_new)  # 路段时间
                path_time_new = self.PathTime(LinkTime_new)  # 路径时间
                od_flow_new = self.ODFlow(pathflow_new)  # OD的角色流量
                total_cost_new = self.TotalCost(path_time_new, od_flow_new, VOT, COI, benchmark, surge_cof, sundry)  # 总成本
                lambda_new = self.Multiplier(total_cost_new)  # 乘子
                general_cost_new = self.GeneralCost(total_cost_new, lambda_new)  # 广义出行成本
                break
            else:
                pathflow_old = copy.deepcopy(pathflow_new)  # 更新路径流量，继续迭代
                LinkTime_old = self.LinkTime(pathflow_old)  # 路段时间
                path_time_old = self.PathTime(LinkTime_old)  # 路径时间
                od_flow_old = self.ODFlow(pathflow_old)  # OD的角色流量
                total_cost_old = self.TotalCost(path_time_old, od_flow_old, VOT, COI, benchmark, surge_cof, sundry)  # 总成本
                lambda_old = self.Multiplier(total_cost_old)  # 乘子
                general_cost_old = self.GeneralCost(total_cost_old, lambda_old)  # 广义出行成本

        print("MSA迭代次数 = ", iter)
        time_2 = time.time()
        print("MSA总共用时：", time_2 - time_1, "s = ", (time_2 - time_1) / 60, "min")
        return [pathflow_new, general_cost_new]


    def CheckResult(self, pathflow, general_cost):   # 检查最终结果
        Pi = mat(np.zeros([self.Route_all, self.Role_Num]))
        for od_i in range(self.OD_Num):  # OD对
            for i in range(self.Route_Num[od_i]):   # 路线 route
                for j in range(self.Role_Num):  # 角色数量
                    if pathflow[(self.Route_Start[od_i] + i) * self.Role_Num + j, 0] > 0:   # 当 alternative 的流量为 0 时
                        Pi[self.Route_Start[od_i] + i, j] = general_cost[(self.Route_Start[od_i] + i) * self.Role_Num + j, 0]
        return Pi


    def SaveNetInfo(self, filename):   # 保存网络信息
        wb = load_workbook(filename)
        # 配置字体格式为：样式（Cambria）、尺寸（11）、斜体（flase）、颜色（黑色）、粗体（false），这里可以根据自己需求修改
        font_style = Font(name='Cambria', size=11, italic=False, color='FF000000', bold=False)
        print("开始写入网络信息")

        # 将 OD_Route 写入文件中
        ws0 = wb.create_sheet("OD_Route")
        for od_i in range(Net.OD_Num):
            temp_od_route = Net.OD_Route[od_i].tolist()
            ws0.append(temp_od_route)
            for j in range(Net.Route_all):
                ws0.cell(row=od_i + 1, column=j + 1).font = font_style

        # 将 Route_Link 写入文件中
        ws1 = wb.create_sheet("Route_Link")
        for i in range(Net.Route_all):
            temp_route_link = Net.Route_Link[i].tolist()
            ws1.append(temp_route_link)
            for j in range(Net.Route_all):
                ws1.cell(row=i + 1, column=j + 1).font = font_style

        wb.save(filename)
        print("网络信息已保存")


    def SaveResult(self, pathflow, general_cost, Pi, rider_Ca, filename):   # 保存计算结果
        wb = load_workbook(filename)
        # 配置字体格式为：样式（Cambria）、尺寸（11）、斜体（flase）、颜色（黑色）、粗体（false），这里可以根据自己需求修改
        font_style = Font(name='Cambria', size=11, italic=False, color='FF000000', bold=False)
        print("开始写入计算结果")

        # 将 pathflow 结果写入文件中
        ws2 = wb.create_sheet("pathflow")
        for od_i in range(Net.OD_Num):
            header = ["SD"]
            for n in range(rider_Ca):
                header.append("RD" + str(n + 1))
            for n in range(rider_Ca):
                header.append("R_" + str(n + 1))
            ws2.append(header)
            for i in range(Net.Route_Num[od_i]):
                temp_od_pathflow = []
                for j in range(Net.Role_Num):
                    temp_od_pathflow.append(pathflow[(Net.Route_Start[od_i] + i) * Net.Role_Num + j, 0])
                ws2.append(temp_od_pathflow)
                for k in range(Net.Role_Num):
                    ws2.cell(row=Net.Route_Start[od_i] + i + 1, column=k + 1).font = font_style

        # 将 general_cost 结果写入文件中
        ws3 = wb.create_sheet("general_cost")
        for od_i in range(Net.OD_Num):
            ws3.append(header)
            for i in range(Net.Route_Num[od_i]):
                temp_od_gen_cost = []
                for j in range(Net.Role_Num):
                    temp_od_gen_cost.append(general_cost[(Net.Route_Start[od_i] + i) * Net.Role_Num + j, 0])
                ws3.append(temp_od_gen_cost)
                for k in range(Net.Role_Num):
                    ws3.cell(row=Net.Route_Start[od_i] + i + 1, column=k + 1).font = font_style

        # 将 π 结果写入文件中
        ws4 = wb.create_sheet("Multiplier")
        for od_i in range(Net.OD_Num):
            ws4.append(header)
            for i in range(Net.Route_Num[od_i]):
                temp_Pi = []
                for j in range(Net.Role_Num):
                    temp_Pi.append(Pi[Net.Route_Start[od_i] + i, j])
                ws4.append(temp_Pi)

        wb.save(filename)
        print("计算结果已保存")


#############测试代码#############

# # ##### ################## #####
# filename = "E:\\课题组\\2022实践创新项目\\MSA求解RUE\\Sioux-Falls.xlsx"
# rider_Ca = 2
# K = 3
# accuracy_0 = 0.01
# VOT = mat(np.array([[1], [0.8], [0.8], [0.4], [0.3]]))
# COI = mat(np.array([[0], [0.3], [0.4], [0.3], [0.4]]))
# benchmark = mat(np.array([[0], [20], [20], [20], [20]]))
# surge_cof = mat(np.array([[0], [5], [5], [1], [1]]))
# sundry = mat(np.array([[1], [1], [1], [0], [0]]))
# # ##### ################## #####

Net = TransNet()

# Net.InitialNet(filename, rider_Ca, K)   # Initialize network topology
# pathflow_0 = Net.Initial_PathFlow()  # Initialize the path flow
# Net.SaveNetInfo(filename)   # Save the route information: OD_Route matrix and Route_Link matrix
#
# # ### Solve path flow and generalized travel cost through the CA algorithm
# [pathflow_CA, general_cost_CA] = Net.CA(pathflow_0, VOT, COI, benchmark, surge_cof, sundry, accuracy_0)
# Pi_CA = Net.CheckResult(pathflow_CA, general_cost_CA)  # Calculate the Lagrangian multiplier
# Net.SaveResult(pathflow_CA, general_cost_CA, Pi_CA, rider_Ca, filename)  # Save the path flow, travel cost and multiplier results
#
# # ### Solve path flow and generalized travel cost through the MSA algorithm
# [pathflow_MSA, general_cost_MSA] = Net.MSA(pathflow_0, VOT, COI, benchmark, surge_cof, sundry, accuracy_0)
# Pi_MSA = Net.CheckResult(pathflow_MSA, general_cost_MSA)  # Calculate the Lagrangian multiplier
# Net.SaveResult(pathflow_MSA, general_cost_MSA, Pi_MSA, rider_Ca, filename)  # Save the path flow, travel cost and multiplier results




