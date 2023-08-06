from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter

import json
import yaml


class Checkxtractor(object):
    """ Extract validation value from xlsx file (only) """
    def __init__(self, source, output, sheet=0, row=0, template_type="python"):
        self.source = source
        self.output = output
        self.sheet = int(sheet)
        self.row = int(row)
        self.template_type = template_type
        self.columns_list = []
        self.validation_list = []
        self.names = {}
        self.wb = None
        self.ws = None
        self.set_values = {}
        self.used_validators = set()

    def extract(self):
        self.wb = load_workbook(self.source)
        for name in self.wb.defined_names.definedName:
            if name.destinations:
                self.names[name.name] = name.destinations
        self.ws = self.wb.worksheets[self.sheet]
        self.columns_list = []
        validation_dict = {}
        # Get active columns with name
        for cell in self.ws[self.row + 1]:
            if cell.value:
                self.columns_list.append(cell.value)
        if not self.columns_list:
            raise Exception("TODO")
        validation_order = {}
        # Need to re-order the keys
        for validation in self.ws.data_validations.dataValidation:
            if validation.type is None:
                continue
            for cell_range in validation.sqref.ranges:
                associated_columns = self._get_column(cell_range)
                for col in associated_columns:
                    if col > len(self.columns_list):
                        continue
                    validation_order[col] = validation
        for col, validation in sorted(validation_order.items()):
            predicted_type = self._predict_type(validation, get_column_letter(col))
            # Will be overriden if conflicting values...
            validation_dict[self.columns_list[col - 1]] = predicted_type
        self._generate_script(self.output, validation_dict)

    def _get_column(self, cell_range):
        if cell_range.min_row > self.row + 2:
            return []

        if cell_range.min_row == cell_range.max_row and not (cell_range.min_col == cell_range.max_col):
            # Might be a mistake, ignore it
            return []
        return set([cell_range.min_col, cell_range.max_col])

    def _predict_type(self, validation, column_name):
        if validation.type == "decimal":
            self.used_validators.add("FloatValidator")
            return self._format_validator("FloatValidator", self._get_numbers_limits(validation, is_float=True))
        if validation.type == "whole":
            self.used_validators.add("IntValidator")
            return self._format_validator("IntValidator", self._get_numbers_limits(validation))
        if validation.type == "date":
            self.used_validators.add("DateValidator")
            return self._format_validator("DateValidator", {})
        if validation.type == "list":
            if not validation.formula1:
                return self._format_validator("NoValidator", {})
            if validation.formula1.startswith("INDIRECT("):
                self.used_validators.add("LinkedSetValidator")
                return self._format_validator("LinkedSetValidator", self._get_linked_set_values(validation, column_name))
            else:
                self.used_validators.add("SetValidator")
                return self._format_validator("SetValidator", self._get_set_values(validation, column_name))
        else:
            return self._format_validator("NoValidator", {})

    def _get_numbers_limits(self, validation, is_float=False):
        cast_func = float if is_float else int
        # Compatibility for "between"
        if validation.operator == "between" or (validation.operator is None and validation.formula1 is not None and validation.formula2 is not None):
            return {"min": cast_func(validation.formula1), "max": cast_func(validation.formula2)}
        elif validation.operator in ["greaterThan", "greaterThanOrEqual"]:
            return {"min": cast_func(validation.formula1)}
        elif validation.operator in ["lessThan", "lessThanOrEqual"]:
            return {"max": cast_func(validation.formula1)}
        else:
            return {}

    def _get_linked_set_values(self, validation, column_name):
        formula = validation.formula1
        values_dict = {}
        cell_coord = formula.split("INDIRECT($")[-1].split(")")[0]
        related_column_name = self.columns_list[self.ws[cell_coord].column - 1]
        cell_column = self.ws[cell_coord].column_letter
        if cell_column not in self.set_values:
            return {}
        for value in self.set_values[cell_column]:
            if value in self.names:
                values_dict[value] = []
                for sheet, coords in self.names[value]:
                    ws = self.wb[sheet]
                    if isinstance(ws[coords], tuple):
                        for cell_tuple in ws[coords]:
                            for cell in cell_tuple:
                                values_dict[value].append(cell.value)
                    else:
                        values_dict[value].append(ws[coords].value)
        if not values_dict:
            return {}
        return {"linked_column": related_column_name, "valid_values": values_dict}

    def _get_set_values(self, validation, column_name):
        formula = validation.formula1
        if "," in formula:
            value_list = formula.replace('"', '').split(",")
            self.set_values[column_name] = value_list
            return {'valid_values': value_list}
        # If it uses names, extract values
        elif formula.lstrip("=") in self.names:
            cell_range = self.names[formula.lstrip("=")]
            value_list = []
            for sheet, coords in cell_range:
                ws = self.wb[sheet]
                if isinstance(ws[coords], tuple):
                    for cell_tuple in ws[coords]:
                        for cell in cell_tuple:
                            value_list.append(cell.value)
                else:
                    value_list.append(ws[coords].value)
            self.set_values[column_name] = value_list
            return {'valid_values': value_list}
        try:
            cell_range = CellRange(range_string=formula.lstrip("="))
            value_list = []
            ws = self.ws
            if cell_range.title:
                ws = self.wb[cell_range.title]
            for cell_coord in cell_range.cells:
                coord = "{}{}".format(get_column_letter(cell_coord[1]), cell_coord[0])
                value_list.append(ws[coord].value)
            self.set_values[column_name] = value_list
            return {'valid_values': value_list}
        except ValueError:
            return {}
        else:
            return {}

    def _format_validator(self, validator, options={}):
        if self.template_type in ['json', 'yml']:
            data = {'type': validator}
            if options:
                data['options'] = options
            return data
        else:
            return "{}({})".format(validator, self._stringify_dict(options))

    def _stringify_dict(self, data):
        if not data:
            return ""
        list_data = []
        for key, values in data.items():
            list_data.append("{}={}".format(key, repr(values)))
        return ", ".join(list_data)

    def _generate_script(self, output_file, validation_dict):
        if self.template_type == "python":
            data = self._generate_python_script(validation_dict)
            with open(output_file, 'w') as f:
                f.write(data)
        elif self.template_type in ['json', 'yml', 'yaml']:
            data = self._generate_template_script(validation_dict)
            with open(output_file, 'w') as f:
                if self.template_type == "json":
                    json.dump(data, f, ensure_ascii=False, indent=4)
                else:
                    f.write(yaml.dump(data, default_flow_style=False))

    def _generate_python_script(self, validation_dict):
        self.used_validators.add("NoValidator")
        validators_list = list(self.used_validators)
        content = ("from checkcel import Checkplate\n"
                   "from checkcel.validators import {}\n"
                   "from collections import OrderedDict\n"
                   "\n"
                   "\n"
                   "class MyTemplate(Checkplate):\n"
                   "    validators = OrderedDict([\n"
                   ).format(", ".join(validators_list))
        for column in self.columns_list:
            validator = validation_dict.get(column, "NoValidator()")
            content += '        ("{}", {}),\n'.format(column, validator)
        content = content.rstrip(",\n") + "\n"
        content += "    ])\n"
        return content

    def _generate_template_script(self, validation_dict):
        data = {"empty_ok": False, "validators": []}
        for column in self.columns_list:
            validator = validation_dict.get(column, {"type": "NoValidator"})
            validator['name'] = column
            data['validators'].append(validator)

        return data
